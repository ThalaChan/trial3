import os
import random
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import GRID_UNITS, STEP, DRONE_FIELDS, DEFAULT_DRONE_SPEED

"""
Excel / CSV Import
==================
Reads drone configuration from .xlsx or .csv.

Column names are driven by DRONE_FIELDS in config.py plus the fixed
coordinate columns (src_x, src_y, dst_x, dst_y).

To support a new drone property in the import file:
  1. Add it to DRONE_FIELDS in config.py
  2. Done — file_loader reads DRONE_FIELDS automatically.
"""

VALID_COORDS = set(range(0, GRID_UNITS + 1, STEP))


def _snap(val):
    return round(float(val) / STEP) * STEP


def _rand_node():
    nodes = list(VALID_COORDS)
    return (random.choice(nodes), random.choice(nodes))


def _parse_coord(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "none", "null", "nan", "-", "n/a"):
        return None
    try:
        v = int(_snap(float(s)))
        return v if v in VALID_COORDS else None
    except (ValueError, TypeError):
        return None


def _row_to_config(row: dict) -> dict:
    """
    Convert one normalised row dict to a drone config dict.
    Coordinates are handled separately; all other fields come from
    DRONE_FIELDS so this function never needs editing.
    """
    # Source
    sx = _parse_coord(row.get("src_x"))
    sy = _parse_coord(row.get("src_y"))
    if sx is None or sy is None:
        rnd = _rand_node()
        sx  = sx if sx is not None else rnd[0]
        sy  = sy if sy is not None else rnd[1]
    src = (sx, sy)

    # Destination
    dx = _parse_coord(row.get("dst_x"))
    dy = _parse_coord(row.get("dst_y"))
    if dx is None or dy is None:
        rnd = _rand_node()
        dx  = dx if dx is not None else rnd[0]
        dy  = dy if dy is not None else rnd[1]
    dst = (dx, dy)

    attempts = 0
    while dst == src and attempts < 20:
        dst = _rand_node()
        attempts += 1

    cfg = {"src": src, "dst": dst}

    # All other fields from DRONE_FIELDS
    for field in DRONE_FIELDS:
        raw = row.get(field["name"].lower())
        val = field["parse"](raw) if raw is not None else None
        cfg[field["drone_attr"]] = (val if val is not None
                                    else field["default"])

    return cfg


def _load_xlsx(filepath: str) -> list:
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl required for Excel import — pip install openpyxl")

    wb   = openpyxl.load_workbook(filepath, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [str(h).strip().lower() if h is not None else ""
               for h in rows[0]]
    configs = []
    for raw_row in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        row = {headers[i]: raw_row[i]
               for i in range(min(len(headers), len(raw_row)))}
        configs.append(_row_to_config(row))
    return configs


def _load_csv(filepath: str) -> list:
    import csv
    configs = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        sample = f.read(2048); f.seek(0)
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        for raw_row in reader:
            row = {k.strip().lower(): v for k,v in raw_row.items()}
            configs.append(_row_to_config(row))
    return configs


def load_file(filepath: str) -> list:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        configs = _load_xlsx(filepath)
    elif ext in (".csv", ".txt"):
        configs = _load_csv(filepath)
    else:
        raise ValueError(f"Unsupported file type '{ext}'. Use .xlsx or .csv")
    if not configs:
        raise ValueError("File has no data rows.")
    return configs


def validate_file(filepath: str) -> tuple:
    try:
        configs = load_file(filepath)
        speeds  = [c.get("speed_mult", DEFAULT_DRONE_SPEED)
                   for c in configs]
        msg = (f"Valid — {len(configs)} drone(s). "
               f"Speed range: {min(speeds):.1f}×–{max(speeds):.1f}×")
        return True, msg, configs[:5]
    except Exception as e:
        return False, str(e), []


def create_template(filepath: str = None) -> str:
    """
    Write a blank template Excel with columns driven by DRONE_FIELDS.
    Adding a new field to DRONE_FIELDS automatically adds a column here.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("pip install openpyxl")

    if filepath is None:
        filepath = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "drone_template.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drones"

    hdr_fill  = PatternFill("solid", fgColor="1E2A4A")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="C0C8D8")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fixed coordinate columns + all DRONE_FIELDS columns
    all_cols = ["src_x", "src_y", "dst_x", "dst_y"] + \
               [f["name"] for f in DRONE_FIELDS]

    for ci, h in enumerate(all_cols, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = hdr_align; c.border = border
        ws.column_dimensions[c.column_letter].width = max(len(h)+4, 10)
    ws.row_dimensions[1].height = 22

    # Sample rows
    sample_fill = PatternFill("solid", fgColor="EEF2FF")
    blank_fill  = PatternFill("solid", fgColor="F7F9FF")
    center      = Alignment(horizontal="center", vertical="center")
    data_font   = Font(size=10)

    sample_rows = [
        [0,   0,   50, 50, 1.5],
        [50,  0,   0,  50, 1.0],
        [0,   25,  50, 25, 0.8],
        [None,None,30, 40, None],
        [None,None,None,None,None],
    ]
    for ri, row_data in enumerate(sample_rows, 2):
        # Pad row if DRONE_FIELDS has more than one field
        while len(row_data) < len(all_cols):
            row_data.append(None)
        for ci, val in enumerate(row_data[:len(all_cols)], 1):
            c = ws.cell(row=ri, column=ci,
                        value=val if val is not None else "")
            c.fill = sample_fill if ri%2==0 else blank_fill
            c.font = data_font; c.alignment = center; c.border = border
        ws.row_dimensions[ri].height = 18

    # Instructions sheet
    wi = wb.create_sheet("Instructions")
    wi.column_dimensions["A"].width = 60
    lines = [
        ("UTM Simulator — Excel Import Template", True),
        ("", False),
        ("Fill in the Drones sheet. Each row = one drone.", False),
        ("", False),
        ("FIXED COLUMNS:", True),
        ("  src_x, src_y — source grid coordinates (0–50, multiples of 5)", False),
        ("  dst_x, dst_y — destination grid coordinates", False),
        ("", False),
        ("DYNAMIC COLUMNS (from DRONE_FIELDS in config.py):", True),
    ]
    for field in DRONE_FIELDS:
        display = field.get("display") or field["name"]
        lines.append((f"  {field['name']} — {display} "
                      f"(default: {field['default']})", False))
    lines += [
        ("", False),
        ("BLANK CELL RULES:", True),
        ("  Coordinates  blank → random valid position", False),
        ("  Other fields blank → default value from config.py", False),
    ]
    hdr_fill2 = PatternFill("solid", fgColor="1E2A4A")
    for ri,(text,bold) in enumerate(lines,1):
        c = wi.cell(row=ri, column=1, value=text)
        c.font = Font(bold=bold, size=11 if bold else 10,
                      color="FFFFFF" if ri==1 else "1E2A4A")
        if ri == 1: c.fill = hdr_fill2
        wi.row_dimensions[ri].height = 18

    wb.save(filepath)
    print(f"[Template] {filepath}")
    return filepath