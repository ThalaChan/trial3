import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side)
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[Logger] openpyxl not found — pip install openpyxl")

RESULTS_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "simulation_results")


def _ensure_results_dir():
    os.makedirs(RESULTS_DIR, exist_ok=True)


def _next_run_number() -> int:
    _ensure_results_dir()
    highest = 0
    for fname in os.listdir(RESULTS_DIR):
        stem = os.path.splitext(fname)[0]
        if stem.startswith("trial_run_"):
            try:
                n = int(stem.replace("trial_run_", ""))
                highest = max(highest, n)
            except ValueError:
                pass
    return highest + 1


class CollisionLogger:

    _COL_HEADER_BG = "1E2A4A"
    _COL_HEADER_FG = "FFFFFF"
    _COL_BORDER    = "C0C8D8"

    # Row fill colors
    _FILL_ARRIVED   = "D6F4DD"   # green
    _FILL_DIRECT    = "FFE0E0"   # red
    _FILL_PROX      = "E0EEFF"   # blue
    _FILL_BATTERY   = "E8D5FF"   # purple
    _FILL_CANCELLED = "E8E8E8"   # gray
    _FILL_NEARMISS  = "FFF8E0"   # yellow
    _FILL_ACTIVE    = "F7F7F7"   # light gray

    def __init__(self):
        self._events          = []
        self._event_id        = 0
        self._run             = 0
        self._run_label       = ""
        self._drones          = []
        self._nearmiss_active = set()

    def next_run(self, label: str = ""):
        self._run      += 1
        self._run_label = label if label else f"Run {self._run}"
        self._nearmiss_active.clear()

    def log(self, grid_tick: int, kind: str,
            severity: str, drone_a, drone_b):
        if kind == "near_miss":
            pair = (min(drone_a.did, drone_b.did),
                    max(drone_a.did, drone_b.did))
            if pair in self._nearmiss_active:
                return
            self._nearmiss_active.add(pair)
        else:
            pair = (min(drone_a.did, drone_b.did),
                    max(drone_a.did, drone_b.did))
            self._nearmiss_active.discard(pair)

        self._event_id += 1

        from config import LOG_COLUMNS
        event = {}
        for col in LOG_COLUMNS:
            h = col["header"]
            if h == "Event ID":
                event[h] = self._event_id
            elif h == "Path Run":
                event[h] = self._run
            elif h == "Path Label":
                event[h] = self._run_label
            elif h == "Timestamp":
                event[h] = datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S.%f")[:-3]
            elif h == "Type":
                event[h] = kind.replace("_"," ").title()
            elif h == "Severity":
                event[h] = severity
            else:
                try:
                    event[h] = col["value"](
                        grid_tick, drone_a, drone_b)
                except Exception:
                    event[h] = ""
        self._events.append(event)

    def save(self, custom_name: str = "",
             drones=None) -> str:
        _ensure_results_dir()
        self._drones = drones or []

        if custom_name.strip():
            stem = (custom_name.strip()
                    .replace("/","_").replace("\\","_")
                    .replace(":","_").replace("*","_")
                    .replace("?","_").replace('"',"_")
                    .replace("<","_").replace(">","_")
                    .replace("|","_"))
        else:
            stem = f"trial_run_{_next_run_number()}"

        ext      = "xlsx" if OPENPYXL_AVAILABLE else "csv"
        filepath = os.path.join(RESULTS_DIR, f"{stem}.{ext}")
        counter  = 1
        while os.path.exists(filepath):
            filepath = os.path.join(
                RESULTS_DIR, f"{stem}_{counter}.{ext}")
            counter += 1

        if OPENPYXL_AVAILABLE:
            self._save_xlsx(filepath)
        else:
            self._save_csv(filepath)

        print(f"[Logger] {len(self._events)} event(s) → {filepath}")
        return filepath

    def reset(self):
        self._events          = []
        self._event_id        = 0
        self._run             = 0
        self._run_label       = ""
        self._drones          = []
        self._nearmiss_active = set()

    def total(self) -> int:
        return len(self._events)

    def summary(self) -> dict:
        direct   = sum(1 for e in self._events
                       if e.get("Type","").lower()=="direct")
        prox     = sum(1 for e in self._events
                       if e.get("Type","").lower()=="proximity")
        nearmiss = sum(1 for e in self._events
                       if "near" in e.get("Type","").lower())
        return {"total"    : len(self._events),
                "direct"   : direct,
                "proximity": prox,
                "near_miss": nearmiss}

    # ── styles ────────────────────────────────────────────────────────────────

    def _border(self):
        s = Side(style="thin", color=self._COL_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr(self):
        return (
            PatternFill("solid", fgColor=self._COL_HEADER_BG),
            Font(bold=True, color=self._COL_HEADER_FG, size=11),
            Alignment(horizontal="center", vertical="center"),
        )

    def _fill(self, hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _status_style(self, status: str):
        """Returns (fill, font_color) for a drone status."""
        m = {
            "arrived"            : (self._FILL_ARRIVED,   "1A7A3A"),
            "crashed_direct"     : (self._FILL_DIRECT,    "A32D2D"),
            "crashed_prox"       : (self._FILL_PROX,      "185FA5"),
            "incomplete_battery" : (self._FILL_BATTERY,   "4B0082"),
            "cancelled"          : (self._FILL_CANCELLED, "555555"),
        }
        bg, fg = m.get(status, (self._FILL_ACTIVE, "444444"))
        return self._fill(bg), Font(size=10, color=fg, bold=True)

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _save_xlsx(self, filepath):
        wb  = openpyxl.Workbook()
        ws1 = wb.active;        ws1.title = "Drone Summary"
        self._write_drone_summary_sheet(ws1)
        ws2 = wb.create_sheet("Collision Log")
        self._write_collision_log_sheet(ws2)
        ws3 = wb.create_sheet("Summary")
        self._write_summary_sheet(ws3)
        ws4 = wb.create_sheet("By Path Run")
        self._write_by_run_sheet(ws4)
        wb.save(filepath)

    # ── Sheet 1: Drone Summary (first sheet, most important) ──────────────────

    def _write_drone_summary_sheet(self, ws):
        """
        One row per drone.
        Columns (exactly as requested):
          Drone | Start Time | End Time | Flight Status |
          Collision Severity | Vehicle | Path Label |
          Drone Speed | Drone Source | Drone Destination |
          Drone Coord X | Drone Coord Y |
          Crashed With Drone |
          Battery at Start | Battery at End | Battery Consumed |
          Distance Planned | Distance Actual
        """
        fill_h, font_h, align_h = self._hdr()
        border  = self._border()
        center  = Alignment(horizontal="center", vertical="center")
        left    = Alignment(horizontal="left",   vertical="center")
        dfont   = Font(size=10)

        headers = [
            "Drone",
            "Start Time",
            "End Time",
            "Flight Status",
            "Collision Severity",
            "Vehicle",
            "Path Label",
            "Drone Speed",
            "Drone Source",
            "Drone Destination",
            "Drone Coord X",
            "Drone Coord Y",
            "Crashed With Drone",
            "Battery at Start",
            "Battery at End",
            "Battery Consumed",
            "Distance Planned",
            "Distance Actual",
        ]

        # Build collision lookup: drone_id → event
        collision_map  = {}
        nearmiss_map   = {}
        for e in self._events:
            kind = e.get("Type","").lower()
            for side in ("A", "B"):
                did = e.get(f"Drone {side} ID")
                if did is None:
                    continue
                if kind != "near miss":
                    if did not in collision_map:
                        collision_map[did] = e
                else:
                    if did not in nearmiss_map:
                        nearmiss_map[did] = e

        # Title
        last = ws.cell(row=1, column=len(headers)).column_letter
        ws.merge_cells(f"A1:{last}1")
        t           = ws["A1"]
        t.value     = "Drone Summary"
        t.font      = Font(bold=True,
                           color=self._COL_HEADER_FG, size=13)
        t.fill      = self._fill(self._COL_HEADER_BG)
        t.alignment = align_h
        ws.row_dimensions[1].height = 26

        # Header row
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.fill=fill_h; c.font=font_h
            c.alignment=align_h; c.border=border
        ws.row_dimensions[2].height = 20

        drones = sorted(getattr(self,"_drones",[]),
                        key=lambda d: d.did)
        if not drones:
            ws.cell(row=3, column=1,
                    value="No data — export after simulation")
            return

        for ri, d in enumerate(drones, 3):
            row_fill, sfont = self._status_style(d.status)

            # Collision event for this drone
            ev   = collision_map.get(d.did, {})
            nm_ev= nearmiss_map.get(d.did, {})

            # Who did this drone crash with?
            crashed_with = "—"
            sev          = "—"
            coord_x      = "—"
            coord_y      = "—"
            if ev:
                if ev.get("Drone A ID") == d.did:
                    crashed_with = ev.get("Drone B ID", "—")
                    coord_x      = ev.get("Drone A Coord X", "—")
                    coord_y      = ev.get("Drone A Coord Y", "—")
                else:
                    crashed_with = ev.get("Drone A ID", "—")
                    coord_x      = ev.get("Drone B Coord X", "—")
                    coord_y      = ev.get("Drone B Coord Y", "—")
                sev = ev.get("Severity", "—")

            bat_start = getattr(d, "battery_start",   100.0)
            bat_end   = getattr(d, "battery_current", 100.0)
            bat_used  = bat_start - bat_end
            spd       = getattr(d, "speed_mult",      1.0)
            vtype     = getattr(d, "vehicle_type",    "quad")
            dist_plan = round(getattr(d, "planned_distance",d.nav.path_cost()), 2)
            dist_act  = round(getattr(d, "actual_distance", 0.0), 2)

            row_data = [
                d.did,
                getattr(d, "start_time", "—"),
                getattr(d, "end_time",   "—"),
                getattr(d, "flight_status", "—"),
                sev,
                vtype,
                d.nav.rank_label(),
                f"{spd:.2f}×",
                str(d.src),
                str(d.dst),
                coord_x,
                coord_y,
                crashed_with,
                f"{bat_start:.1f}%",
                f"{bat_end:.1f}%",
                f"{bat_used:.1f}%",
                dist_plan,
                dist_act,
            ]

            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill      = row_fill
                c.border    = border
                c.alignment = center
                # Status column bold colored font
                c.font = sfont if ci == 4 else dfont
            ws.row_dimensions[ri].height = 18

        # Auto-fit
        col_w = [len(h) for h in headers]
        for d in drones:
            sample = [str(d.did),
                      "2026-05-18 14:00:00.000",
                      "2026-05-18 14:00:00.000",
                      "Collision — Node to Node",
                      "Critical","fixed_wing",
                      "#1 optimal","1.50×",
                      str(d.src),str(d.dst),
                      "50","50","99",
                      "100.0%","0.0%","100.0%",
                      "99.99","99.99"]
            for i,v in enumerate(sample):
                col_w[i] = max(col_w[i], len(v))
        for ci,w in enumerate(col_w,1):
            letter = ws.cell(row=2,column=ci).column_letter
            ws.column_dimensions[letter].width = min(w+3,40)

        ws.freeze_panes    = "A3"
        ws.auto_filter.ref = f"A2:{last}2"

        # Legend
        leg = len(drones)+4
        ws.merge_cells(f"A{leg}:F{leg}")
        lc = ws.cell(row=leg,column=1,value="Legend")
        lc.font      = Font(bold=True,size=10,
                            color=self._COL_HEADER_FG)
        lc.fill      = self._fill(self._COL_HEADER_BG)
        lc.alignment = left
        ws.row_dimensions[leg].height=18

        items = [
            ("Complete ✓",             self._FILL_ARRIVED,   "1A7A3A"),
            ("Collision — Node to Node ✖", self._FILL_DIRECT,    "A32D2D"),
            ("Collision — Proximity ⚠", self._FILL_PROX,      "185FA5"),
            ("Incomplete — Battery ⚡", self._FILL_BATTERY,   "4B0082"),
            ("Cancelled ✕",            self._FILL_CANCELLED, "555555"),
            ("Active / Waiting",       self._FILL_ACTIVE,    "444444"),
        ]
        for i,(lbl,bg,fg) in enumerate(items):
            r = leg+1+i
            ws.merge_cells(f"A{r}:F{r}")
            c = ws.cell(row=r,column=1,value=lbl)
            c.fill      = self._fill(bg)
            c.font      = Font(size=10,color=fg,bold=True)
            c.border    = border
            c.alignment = left
            ws.row_dimensions[r].height=16

    # ── Sheet 2: Collision Log ────────────────────────────────────────────────

    def _write_collision_log_sheet(self, ws):
        from config import LOG_COLUMNS
        fill_h, font_h, align_h = self._hdr()
        border = self._border()
        center = Alignment(horizontal="center",vertical="center")
        dfont  = Font(size=10)

        cols = [c["header"] for c in LOG_COLUMNS]

        for ci,col in enumerate(cols,1):
            c = ws.cell(row=1,column=ci,value=col)
            c.fill=fill_h; c.font=font_h
            c.alignment=align_h; c.border=border
        ws.row_dimensions[1].height=22

        for ri,event in enumerate(self._events,2):
            kind = event.get("Type","").lower()
            if "direct" in kind:
                rf = self._fill(self._FILL_DIRECT)
            elif "proximity" in kind:
                rf = self._fill(self._FILL_PROX)
            else:
                rf = self._fill(self._FILL_NEARMISS)
            for ci,col in enumerate(cols,1):
                c = ws.cell(row=ri,column=ci,
                            value=event.get(col,""))
                c.fill=rf; c.font=dfont
                c.alignment=center; c.border=border
            ws.row_dimensions[ri].height=18

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

        widths = {col:len(col) for col in cols}
        for e in self._events:
            for col in cols:
                widths[col]=max(widths[col],
                                len(str(e.get(col,""))))
        for ci,col in enumerate(cols,1):
            letter=ws.cell(row=1,column=ci).column_letter
            ws.column_dimensions[letter].width=min(widths[col]+4,40)

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────

    def _write_summary_sheet(self, ws):
        fill_h, font_h, align_h = self._hdr()
        border  = self._border()
        bfont   = Font(bold=True,size=11)
        dfont   = Font(size=11)
        left    = Alignment(horizontal="left",  vertical="center")
        center  = Alignment(horizontal="center",vertical="center")

        ws.merge_cells("A1:B1")
        t = ws["A1"]
        t.value     = "UTM Simulation Summary"
        t.font      = Font(bold=True,color=self._COL_HEADER_FG,
                           size=13)
        t.fill      = self._fill(self._COL_HEADER_BG)
        t.alignment = align_h
        ws.row_dimensions[1].height=28

        s      = self.summary()
        drones = getattr(self,"_drones",[])
        arrived   = sum(1 for d in drones if d.status=="arrived")
        inc_bat   = sum(1 for d in drones
                        if d.status=="incomplete_battery")
        cancelled = sum(1 for d in drones
                        if d.status=="cancelled")

        rows=[
            ("Total drones",          len(drones),
             self._fill(self._FILL_ACTIVE)),
            ("Complete (arrived)",    arrived,
             self._fill(self._FILL_ARRIVED)),
            ("Incomplete — Battery",  inc_bat,
             self._fill(self._FILL_BATTERY)),
            ("Cancelled",             cancelled,
             self._fill(self._FILL_CANCELLED)),
            ("Total path runs",       self._run,
             self._fill(self._FILL_ACTIVE)),
            ("Total collision events",s["total"],
             self._fill(self._FILL_ACTIVE)),
            ("Direct collisions",     s["direct"],
             self._fill(self._FILL_DIRECT)),
            ("Proximity collisions",  s["proximity"],
             self._fill(self._FILL_PROX)),
            ("Near misses",           s["near_miss"],
             self._fill(self._FILL_NEARMISS)),
        ]
        for i,(label,value,rfill) in enumerate(rows,2):
            lc=ws.cell(row=i,column=1,value=label)
            vc=ws.cell(row=i,column=2,value=value)
            lc.font=bfont; vc.font=dfont
            lc.alignment=left; vc.alignment=center
            lc.fill=rfill; vc.fill=rfill
            lc.border=border; vc.border=border
            ws.row_dimensions[i].height=22

        ws.column_dimensions["A"].width=38
        ws.column_dimensions["B"].width=26

    # ── Sheet 4: By Path Run ──────────────────────────────────────────────────

    def _write_by_run_sheet(self, ws):
        fill_h, font_h, align_h = self._hdr()
        border = self._border()
        dfont  = Font(size=10)
        center = Alignment(horizontal="center",vertical="center")
        left   = Alignment(horizontal="left",  vertical="center")

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value     = "Collisions by Path Run"
        t.font      = Font(bold=True,color=self._COL_HEADER_FG,
                           size=13)
        t.fill      = self._fill(self._COL_HEADER_BG)
        t.alignment = align_h
        ws.row_dimensions[1].height=26

        for ci,h in enumerate(["Path Run","Path Label",
                                "Total","Direct",
                                "Proximity","Near Miss"],1):
            c=ws.cell(row=2,column=ci,value=h)
            c.fill=fill_h; c.font=font_h
            c.alignment=align_h; c.border=border
        ws.row_dimensions[2].height=20

        runs={}
        for e in self._events:
            rid=e.get("Path Run",1)
            if rid not in runs:
                runs[rid]={"label"    : e.get("Path Label",""),
                           "total"    : 0,
                           "direct"   : 0,
                           "prox"     : 0,
                           "nearmiss" : 0}
            runs[rid]["total"]+=1
            t=e.get("Type","").lower()
            if "direct" in t:    runs[rid]["direct"]  +=1
            elif "prox" in t:    runs[rid]["prox"]    +=1
            else:                runs[rid]["nearmiss"]+=1

        for ri,(rid,data) in enumerate(sorted(runs.items()),3):
            rf=(self._fill(self._FILL_DIRECT)
                if data["direct"]>=data["prox"]
                else self._fill(self._FILL_PROX))
            for ci,val in enumerate(
                    [rid,data["label"],data["total"],
                     data["direct"],data["prox"],
                     data["nearmiss"]],1):
                c=ws.cell(row=ri,column=ci,value=val)
                c.fill=rf; c.font=dfont
                c.alignment=left if ci==2 else center
                c.border=border
            ws.row_dimensions[ri].height=18

        tr=len(runs)+3
        ws.merge_cells(f"A{tr}:B{tr}")
        tc=ws.cell(row=tr,column=1,value="TOTAL")
        tc.font=Font(bold=True,color=self._COL_HEADER_FG,size=11)
        tc.fill=self._fill(self._COL_HEADER_BG)
        tc.alignment=center; tc.border=border
        for ci,val in enumerate([
            sum(d["total"]    for d in runs.values()),
            sum(d["direct"]   for d in runs.values()),
            sum(d["prox"]     for d in runs.values()),
            sum(d["nearmiss"] for d in runs.values()),
        ],3):
            c=ws.cell(row=tr,column=ci,value=val)
            c.font=Font(bold=True,color=self._COL_HEADER_FG,
                        size=11)
            c.fill=self._fill(self._COL_HEADER_BG)
            c.alignment=center; c.border=border

        for col,w in zip("ABCDEF",[12,18,10,10,12,12]):
            ws.column_dimensions[col].width=w

    # ── CSV fallback ──────────────────────────────────────────────────────────

    def _save_csv(self, filepath):
        import csv
        from config import LOG_COLUMNS
        cols = [c["header"] for c in LOG_COLUMNS]
        with open(filepath,"w",newline="") as f:
            writer=csv.DictWriter(f,fieldnames=cols)
            writer.writeheader()
            writer.writerows(self._events)